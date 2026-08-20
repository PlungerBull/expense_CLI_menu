"""Read the spreadsheet into plain rows (openpyxl, lazy optional dependency).

The reader returns reader-agnostic ``RawRow``s so the parse layer never touches
openpyxl and stays unit-testable without a real .xlsx file.
"""

from dataclasses import dataclass

from expense.import_ import mapping


@dataclass(frozen=True)
class RawRow:
    """One physical data row. ``line`` is the 1-based sheet row number."""

    line: int
    cells: list[object]


@dataclass(frozen=True)
class SheetData:
    headers: list[object]
    rows: list[RawRow]


class ImportDependencyError(Exception):
    """openpyxl is not installed."""


class ImportFileError(Exception):
    """The file is missing, unreadable, or lacks the expected sheet."""


def read_workbook(path: str, *, sheet_name: str = mapping.SHEET_NAME) -> SheetData:
    """Load ``sheet_name`` from an .xlsx into header + data rows."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportDependencyError(
            "openpyxl is required for `expense import`. Install it with:\n    pip install openpyxl"
        ) from exc

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError as exc:
        raise ImportFileError(f"File not found: {path}") from exc
    except Exception as exc:  # noqa: BLE001 — openpyxl raises a variety of types
        raise ImportFileError(f"Could not open {path}: {exc}") from exc

    try:
        # Tab names are matched case- and whitespace-insensitively: a sheet
        # exported as `DATA`, `data` or `Data ` is the same sheet, and failing
        # on the casing is a papercut with no upside.
        wanted = sheet_name.strip().casefold()
        actual = next((s for s in wb.sheetnames if s.strip().casefold() == wanted), None)
        if actual is None:
            raise ImportFileError(
                f"Sheet {sheet_name!r} not found. Sheets present: {', '.join(wb.sheetnames)}"
            )
        ws = wb[actual]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            raise ImportFileError(f"Sheet {actual!r} is empty.") from None
        # Deliberately materialized (not streamed). `read_only=True` already keeps
        # openpyxl's cell cache lazy; the remaining win from yielding RawRows would
        # require `wb` to stay open through parse — i.e. read_workbook becoming a
        # context manager so the handle closes on every path (parse can raise
        # ImportFormatError on bad headers). Bounded and cheap at personal scale;
        # true streaming is deferred (polish-backlog §5) rather than left half-done
        # with a handle that leaks on the error path.
        data = [RawRow(line=line, cells=list(values)) for line, values in enumerate(rows_iter, 2)]
    finally:
        wb.close()

    return SheetData(headers=header, rows=data)
